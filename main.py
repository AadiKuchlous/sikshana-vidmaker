#!/home/ubuntu/sikshana-vidmaker/env/bin/python

import os
import sys
import boto3
import openpyxl
import shutil
import re


sys.path.append('/usr/bin')
os.system("export PATH=$PATH:/usr/bin")

def aws_polly(text, data_type):
	client = boto3.client('polly')
	if data_type == "json":
		response = client.synthesize_speech(
			LanguageCode='en-IN',
			OutputFormat='json',
			Text=text,
			SpeechMarkTypes=['word'],
			VoiceId='Raveena'
		)
	elif data_type == "mp3":
		response = client.synthesize_speech(
			LanguageCode='en-IN',
			OutputFormat='mp3',
			Text=text,
			VoiceId='Raveena'
		)

	return(response['AudioStream'])


def polly_audio(text):
	audio_data = aws_polly(text, "mp3")
	return(audio_data)


def polly_json(text):
	json_data = aws_polly(text, "json")
	return(json_data)


# def underline_html(text):
# 	html = ""
# 	first = True
# 	for i in text:
# 		if i != '_':
# 			html += i
# 		else:
# 			if first:
# 				html += '<u>'
# 				first = False
# 			else:
# 				html += '</u>'
# 				first = True
# 	return html

def underline_html(text):
	return text.replace("._", "<u>").replace("_.", "</u>")

def create_images(text, image, story=False):
	no_of_lines = len(text.strip().split('\n'))
	text = underline_html(text)
	text = text.split('\n')
	print('lines:')
	print(text)
	new_text = []
	for line in text:
		new_text.append(line.strip())
	new_text = ' '.join(new_text)
	words = new_text.split(' ')
	print("create_images words:{0}".format(words))
	print("no_of_lines: " + str(no_of_lines))
	print(image)
	images = []
	image_name = "tmp.png"
	if type(image) == type(''):
		image_link = "https://drive.google.com/uc?export=download&id="+image.split('d/')[1].split('/view')[0]
		print("link: "+image_link)
		os.system("/usr/bin/wget '{0}' -O '{1}'".format(image_link, image_name))
	for i in range(len(words)):
		text_html = ''
		if words[i][0] in "-:#\n[":
				continue

		for j in range(len(words)):
			if words[j][0] == "#":
				text_html = text_html.strip()+"&nbsp;"*len(words[j])
				continue

			if words[j][0] == '[':
				text_html += words[j][1:-1]
				continue

			if words[j] == '\n':
				text_html += '<br>'
				continue

			if j == i:
				if words[j][-1] in ":,.":
					text_html += '<span style="color:red;">' + words[j][0:-1] + '</span>' + words[j][-1] + " "
				else:
					text_html += '<span style="color:red;">' + words[j] + '</span> '
			else:
				text_html += words[j] + ' '

		br = '<br>'

		logo = '<img src="logo.png" style="position: absolute; top: 0px; left: 0px;"></img>'
		if type(image) == type(''):
			if story:
				if no_of_lines < 4:
					text = '<h1 style="font-size: {0}rem; margin: 0px"><p style="margin: 0px; text-align:center">{1}</p></h1>'.format(3, text_html)
				else:
					text = '<h1 style="font-size: {0}rem; margin: 0px"><p style="margin: 0px; text-align:center">{1}</p></h1>'.format(1, text_html)
				img = '<div style="height:350px;"><img src="{}" style="padding-top:40px;"></img></div>'.format(image_name)
				html = '<!DOCTYPE html><html><body style="margin:0px"><div id="vid_area" style="height: 720px; width: 1280px;">{0}<div style="height: 360px; display: flex; justify-content: center; align-items: center; padding: 20px; padding-top: 30px;">{1}{2}</div></div></body></html>'.format(img, text, logo)

			else:
				img = '<div style="height:425px"><img src="{}" style="padding-top:30px"></img></div>'.format(image_name)
				text = '<h1 style="font-size: {0}rem; margin: 0px"><p style="margin: 0px; text-align:center">{1}</p></h1>'.format(7-no_of_lines, text_html)
				html = '<!DOCTYPE html><html><body style="margin:0px"><div id="vid_area" style="height: 720px; width: 1280px;">{0}<div style="height: 360px; display: flex; justify-content: center; align-items: center; padding:20px">{1}{2}</div></div></body></html>'.format(img, text, logo)
		else:
			if no_of_lines < 4:

				text = '<div style="height: 720px; display: flex; justify-content: center; align-items: center;"><h1 style="font-size: {0}rem; margin-top: 0px"><p style="margin: 0.5em; text-align:center">{1}</p></h1></div>'.format(3.5, text_html)
			else:
				text = '<div style="height: 720px; display: flex; justify-content: center; align-items: center;"><h1 style="font-size: {0}rem; margin-top: 0px"><p style="margin: 0.5em; text-align:center">{1}</p></h1></div>'.format(3, text_html)
			html = '<!DOCTYPE html><html><body style="margin:0px"><div id="vid_area" style="height: 720px; width: 1280px;">{0}{1}</div></body></html>'.format(text, logo)
		print(text)
		with open("tmp.html", "w") as f:
			f.write(html)

		cmd = "/usr/bin/node pup.js file://{0}/tmp.html images/{1}.jpg".format(os.getcwd(), str(i))
		print("PUP CMD: {}".format(cmd))
		os.system(cmd)

		images.append("images/{}.jpg".format(i))

	return(images)


def concatenate_videos(videos, output_name, tmpdir):
	with open('{}/con.in'.format(tmpdir), 'w') as f:
		print(videos)
		for video in videos:
			f.write("file '{0}/{1}'\n".format(tmpdir, video))
	print(sys.path)
	os.system("/usr/bin/ffmpeg -y -f concat -safe 0 -i {0}/con.in -strict -2 -video_track_timescale 90000 -max_muxing_queue_size 2048 -tune animation -crf 6 '{1}'".format(tmpdir, output_name))


def create_para_vid(speed, i, time_data, images, audio, output_name):
	end_times = []
	time_data_l = 0
	for l in time_data.iter_lines():
		time = int(l.decode().split(",")[0][8:])/(1000) * (1/speed)
		time_data_l +=1
		end_times.append(time)

	end_times = end_times[1:]
	print(end_times)

	with open("ffmp.in", "w") as f:
		prev_time = 0.0
		for j in range(len(images)-1):
			duration = float(end_times[j]) - prev_time
			f.write("file {0}' \n".format(images[j]))
			f.write("duration {} \n".format(duration))
			prev_time = float(end_times[j])

		f.write("file '{0}' \n".format(images[-1]))
		f.write("duration {} \n".format(prev_time+0.5))
	os.system("/usr/bin/ffmpeg -y -i {0} -f concat -i ffmp.in -strict -2 -vsync vfr -pix_fmt yuv420p -vf fps=24 -video_track_timescale 90000 -max_muxing_queue_size 2048 -tune animation -crf 6 {1}nf.mp4".format(audio, output_name))
	with open('blank.in', 'w') as f:
		f.write('\n'.join(["file '{0}'".format(images[-1]), "duration {}".format("1.3")]))
	os.system("/usr/bin/ffmpeg -y -i {0} -f concat -i blank.in -strict -2 -vsync vfr -pix_fmt yuv420p -vf fps=24 -video_track_timescale 90000 -max_muxing_queue_size 2048 -tune animation -crf 6 fill{1}.mp4".format("blank_long.mp3", str(i)))
	concatenate_videos(['{0}nf.mp4'.format(output_name), 'fill{}.mp4'.format(i)], output_name+'.mp4', os.getcwd())


def create_intro_video(sheet):
	image = sheet.cell(row=2, column=3).value
	text = (sheet.cell(row=2, column=2).value).split('\n')
	audio_data = polly_audio('. '.join(text))
	with open('audio_uf.mp3', 'wb') as f:
			f.write(audio_data.read())
	os.system("/usr/bin/ffmpeg -y -i {} -ar 48000 {}".format("audio_uf.mp3", "intro_audio.mp3"))
	with open('intro.in', 'w') as f:
		f.write('\n'.join(["file \'images/{0}.jpg\'".format("intro"), "duration {}".format(5)]))
	headers = ''
	for line in text:
		headers += '<h1 style="font-size:{}rem">{}</h1>'.format(4-(0.5*len(text)), line)
	logo = '<img src="logo.png" style="position: absolute; top: 0px; left: 0px;"></img>'
	image_name = "intro.png"
	if type(image) == type(''):
		image_link = "https://drive.google.com/uc?export=download&id="+image.split('d/')[1].split('/view')[0]
		print("link: "+image_link)
		os.system("/usr/bin/wget '{0}' -O '{1}'".format(image_link, image_name))
		print(headers)
		img = '<div style="height:350px"><img src="{}" style=""></img></div>'.format(image_name)
		html = '<!DOCTYPE html><html><body style="margin:0px"><div id="vid_area" style="height: 720px; width: 1280px; padding:10px">{0}<div style="height: 360px; display: flex; justify-content: center; align-items:flex-start; padding-top:{3}px">{1}</div>{2}</div></body></html>'.format(img, headers, logo, 80-(20*len(text)))
	else:
		html = '<!DOCTYPE html><html><body><div id="vid_area" style="height: 720px; width:1280px; display: flex; flex-direction: column; justify-content: center; align-items: center;">{0}{1}</div></body></html>'.format(logo, headers)
	with open("intro.html", "w") as f:
			f.write(html)
	print(os.getcwd())
	cmd = "/usr/bin/node pup.js file://{}/intro.html images/{}.jpg".format(os.getcwd(), "intro")
	print("PUP CMD INTRO:{}".format(cmd))
	os.system(cmd)
	os.system("/usr/bin/ffmpeg -y -i {0} -f concat -i intro.in -strict -2 -vsync vfr -pix_fmt yuv420p -vf fps=24 -video_track_timescale 90000 -max_muxing_queue_size 2048 -tune animation -crf 6 {1}.mp4".format("intro_audio.mp3", "intronf"))
	with open('blank.in', 'w') as f:
		f.write('\n'.join(["file \'images/{0}.jpg\'".format("intro"), "duration {}".format("2.5")]))
	os.system("/usr/bin/ffmpeg -y -i {0} -f concat -i blank.in -strict -2 -vsync vfr -pix_fmt yuv420p -vf fps=24 -video_track_timescale 90000 -max_muxing_queue_size 2048 -tune animation -crf 6 fill{1}.mp4".format("blank.mp3", "intro"))
	concatenate_videos(["intronf.mp4", "fillintro.mp4"], "intro.mp4", os.getcwd())
	return("intro.mp4")


def read_excel(inpfile, sheet):
	name = inpfile
	wb = openpyxl.load_workbook(name)
	print(wb.sheetnames)
	return(wb[sheet])


def create_vids_from_excel(inpfile, sheet, tmpdir, story, first_slide, last_slide):
	mdir = os.getcwd()
	files = ['pup.js', 'blank.mp3', 'blank_long.mp3', 'logo.png']
	for f in files:
		shutil.copy(f, "{0}/{1}".format(tmpdir, f))
	os.chdir(tmpdir)
	os.mkdir('images')
	sheet = read_excel(inpfile, sheet)
	create_intro_video(sheet)
	normal_videos = []
	slow_videos = []
	split_videos = []
	start = int(first_slide)
	end = int(last_slide)+1
	for i in range(start, end):
		if not sheet.cell(row=i, column=1).value:
			continue
		para = sheet.cell(row=i, column=2).value.strip()
		polly_para = para.replace("._", "").replace("_.", "").replace('*', '.').replace('#', '').replace('__', ' - ').replace('_', '').replace('\n','. ')
		polly_para_split = polly_para.split()
		polly_para = ""
		for word in polly_para_split:
			if word[0] != '[':
				polly_para += word + " "
		print("polly_para: "+polly_para)
		json_data = polly_json(polly_para)
		json_data_slow = polly_json(polly_para)
		audio_data = polly_audio(polly_para)
		with open('audio_uf.mp3', 'wb') as f:
			f.write(audio_data.read())
		os.system("/usr/bin/ffmpeg -y -i {} -ar 48000 {}".format("audio_uf.mp3", "audio.mp3"))
		os.system("/usr/bin/sox audio.mp3 audio_slow.mp3 tempo 0.75")
		split_para = '. '.join(polly_para.split())
		json_data_split = polly_json(split_para)
		audio_data_split = polly_audio(split_para)
		with open('audio_uf.mp3', 'wb') as f:
			f.write(audio_data_split.read())
		os.system("/usr/bin/ffmpeg -y -i {} -ar 48000 {}".format("audio_uf.mp3", "audio_split.mp3"))

		images_text = ' '.join(para.strip().replace('*', '').replace('\n', ' \n ').split(' '))
		print(images_text)
		images = create_images(images_text, sheet.cell(row=i, column=3).value, story)
		print("polly_para: "+polly_para)

		create_para_vid(1, i-start-1, json_data, images, 'audio.mp3', "vid{}".format(str(i-start+1)))
		create_para_vid(0.75, i-2, json_data_slow, images, 'audio_slow.mp3', "vid{}-slow".format(str(i-start+1)))
		create_para_vid(1, i-2, json_data_split, images, 'audio_split.mp3', "vid{}-split".format(str(i-start+1)))

		normal_videos.append("vid{}.mp4".format(str(i-start+1)))
		slow_videos.append("vid{}-slow.mp4".format(str(i-start+1)))
		split_videos.append("vid{}-split.mp4".format(str(i-start+1)))

	os.chdir(os.path.join(mdir, 'videos'))
	# os.system("/usr/bin/mkdir {}".format(output_name.replace(' ', '\ ')))
	os.mkdir(output_name)
	print(output_name)
	os.chdir(os.path.join(os.getcwd(), output_name))
	print(os.getcwd())

	concatenate_videos(["intro.mp4"]+normal_videos, "{}-fast.mp4".format(output_name), tmpdir)
	concatenate_videos(["intro.mp4"]+slow_videos, "{}-medium.mp4".format(output_name), tmpdir)
	concatenate_videos(["intro.mp4"]+split_videos, "{}-slow.mp4".format(output_name), tmpdir)

	# os.system("rm *.mp4")
	# os.system("rm a*.mp3")
	# os.system("rm *.in")
	# os.system("rm *_a*.mp3")
	# os.system("rm *.html")
	# os.system("rm -r images")

#read_excel("input.xlsx")
inpfile = str(sys.argv[1])
sheet_name = str(sys.argv[2])
tmpdir = str(sys.argv[3])
output_name = str(sys.argv[4])
story = str(sys.argv[5])
first_slide = sys.argv[6]
last_slide = sys.argv[7]
print(sys.path)
print(tmpdir)
create_vids_from_excel(inpfile, sheet_name, tmpdir, story, first_slide, last_slide)
